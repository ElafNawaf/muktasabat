#!/usr/bin/env python3
"""
Import property-management Excel exports into Muktasabat (Owners, Buildings, Units, Tenants, Contracts).

Expected files (defaults match common Arabic filenames under Downloads):
  1) "بيانات عمائر" workbook — sheet "العمائر" (stacked building blocks + units),
     sheet "عملاء ادارة الاملاك" (owners / clients — data often has name in col A, client # in col B).
  2) "جدول عمائر" workbook — sheet "الورقة1" (repeated sections: building title + tenant rows).

Usage:
  cd repo && .venv/bin/python scripts/migrate_property_excel.py \\
    --data-xlsx "/path/بيانات عمائر ادارة الاملاك.xlsx" \\
    --schedule-xlsx "/path/جدول عمائر إدارة الأملاك.xlsx"

  DATABASE_URL=sqlite:///./muktasbat.db  # or Postgres URL

Options:
  --dry-run   Parse and print counts only (no DB writes).
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Iterator

# Repo root on path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook  # noqa: E402
from sqlalchemy import select  # noqa: E402

from api.database import SessionLocal  # noqa: E402
from api.models import Building, Contract, Owner, Tenant, Unit  # noqa: E402


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def _to_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _cell_str(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(Decimal(str(v).replace(",", "")))
    except Exception:
        return None


def _phone(v: Any) -> str:
    s = _cell_str(v).replace(" ", "")
    if not s or s in ("0", "00"):
        return "0000000000"
    digits = re.sub(r"\D", "", s)
    if not digits:
        return "0000000000"
    return digits[-20:]


def _national_id(raw: Any, phone: str, name: str) -> str:
    s = re.sub(r"\D", "", _cell_str(raw))
    if 8 <= len(s) <= 20:
        return s[:20]
    h = str(abs(hash(name + phone)))[-10:]
    return f"9{h.zfill(9)}"[:20]


def _payment_cycle(pay_method: str) -> int:
    t = (pay_method or "").strip()
    if "دفعة واحدة" in t or "دفعه واحدة" in t:
        return 12
    if "6" in t and "شهر" in t:
        return 6
    if "دفعتين" in t or "دفعتين" in t:
        return 6
    return 6


def _contract_no(prefix: str) -> str:
    u = uuid.uuid4().hex[:10].upper()
    base = f"{prefix}-{u}"
    return base[:50]


def _is_title_row(row: tuple[Any, ...], max_check: int = 12) -> bool:
    if not row or row[0] is None:
        return False
    if not isinstance(row[0], str):
        return False
    first = row[0].strip()
    if first == "رقم الوحدة" or first == "رقم الشقة":
        return False
    rest = [row[i] for i in range(1, min(len(row), max_check)) if i < len(row)]
    nonempty = [x for x in rest if x is not None and _cell_str(x)]
    return len(nonempty) == 0 and len(first) > 3


@dataclass
class UnitRow:
    unit_number: str
    rent: float
    water: float | None
    tenant_name: str
    pay_method: str
    pay1: date | None
    pay2: date | None
    phone: str
    email: str | None
    electric: str | None
    national_id_raw: str | None = None


def iter_alomair_units_sheet(path: str) -> Iterator[tuple[str, list[UnitRow]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["العمائر"]
    except KeyError:
        wb.close()
        return
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    i = 0
    while i < len(rows):
        row = rows[i]
        if not row or row[0] is None:
            i += 1
            continue
        if _is_title_row(row):
            building = _cell_str(row[0])
            i += 1
            if i >= len(rows):
                break
            hdr = rows[i]
            if not hdr or _cell_str(hdr[0]) != "رقم الوحدة":
                continue
            i += 1
            units: list[UnitRow] = []
            while i < len(rows):
                r = rows[i]
                if not r or r[0] is None:
                    i += 1
                    break
                if _is_title_row(r):
                    break
                if _cell_str(r[0]) == "رقم الوحدة":
                    break
                u0 = r[0]
                if isinstance(u0, str) and u0.strip() == "رقم الوحدة":
                    i += 1
                    break
                num = _cell_str(u0)
                if not num:
                    i += 1
                    continue
                rent = _to_float(r[1]) or 0.0
                water = _to_float(r[2])
                tenant = _cell_str(r[3]) or "—"
                pay_m = _cell_str(r[4])
                d1 = _to_date(r[5]) if len(r) > 5 else None
                d2 = _to_date(r[6]) if len(r) > 6 else None
                phone = _phone(r[7] if len(r) > 7 else None)
                email_s = _cell_str(r[8]) if len(r) > 8 and r[8] else None
                nid = r[9] if len(r) > 9 else None
                elec = _cell_str(r[11]) if len(r) > 11 else None
                if elec in ("", "0", "0.0"):
                    elec = None
                units.append(
                    UnitRow(
                        unit_number=num[:50],
                        rent=rent,
                        water=water,
                        tenant_name=tenant[:150],
                        pay_method=pay_m,
                        pay1=d1,
                        pay2=d2,
                        phone=phone,
                        email=email_s[:120] if email_s else None,
                        electric=elec,
                        national_id_raw=_cell_str(nid) if nid else None,
                    )
                )
                i += 1
            if building and units:
                yield building, units
        else:
            i += 1


@dataclass
class ScheduleRow:
    unit_number: str
    tenant_name: str
    rent: float
    water: float | None
    pay_method: str
    start: date | None
    mid: date | None
    end: date | None
    phone: str
    national_id_raw: str | None
    electric: str | None


def iter_schedule_sheet(path: str) -> Iterator[tuple[str, list[ScheduleRow]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["الورقة1"]
    except KeyError:
        wb.close()
        return
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    building = ""
    i = 0
    while i < len(rows):
        row = rows[i]
        if row and isinstance(row[0], str):
            s = row[0].strip()
            if s.startswith("بيان مستأجري عمارة") or (
                "عمارة" in s and len(s) > 10 and _is_title_row(row)
            ):
                building = re.sub(r"^بيان\s+مستأجري\s+", "", s)
                building = re.sub(r"^بيان\s+", "", building).strip()
        if row and _cell_str(row[0]) == "رقم الشقة":
            i += 1
            block: list[ScheduleRow] = []
            while i < len(rows):
                r = rows[i]
                if not r:
                    i += 1
                    continue
                if _is_title_row(r) or _cell_str(r[0]) == "رقم الشقة":
                    break
                un = _cell_str(r[0])
                if not un or un == "رقم الشقة":
                    i += 1
                    continue
                name = _cell_str(r[1]) if len(r) > 1 else ""
                rent = _to_float(r[2]) if len(r) > 2 else 0.0
                water = _to_float(r[3]) if len(r) > 3 else None
                pay_m = _cell_str(r[4]) if len(r) > 4 else ""
                d0 = _to_date(r[5]) if len(r) > 5 else None
                d1 = _to_date(r[6]) if len(r) > 6 else None
                d2 = _to_date(r[7]) if len(r) > 7 else None
                phone = _phone(r[8] if len(r) > 8 else None)
                nid = r[9] if len(r) > 9 else None
                elec = _cell_str(r[10]) if len(r) > 10 else None
                if not name:
                    i += 1
                    continue
                block.append(
                    ScheduleRow(
                        unit_number=un[:50],
                        tenant_name=name[:150],
                        rent=rent or 0.0,
                        water=water,
                        pay_method=pay_m,
                        start=d0,
                        mid=d1,
                        end=d2,
                        phone=phone,
                        national_id_raw=_cell_str(nid) if nid else None,
                        electric=elec if elec and elec not in ("0",) else None,
                    )
                )
                i += 1
            if building and block:
                yield building, block
            # Avoid spinning on back-to-back "رقم الشقة" rows when no data rows were read.
            if not block:
                while i < len(rows) and rows[i] and _cell_str(rows[i][0]) == "رقم الشقة":
                    i += 1
            continue
        i += 1


def import_clients(path: str) -> list[tuple[str, float | None, str | None, str | None, str | None]]:
    """Returns (name, client_no, phone, email, tax_id) from عملاء ادارة الاملاك."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["عملاء ادارة الاملاك"]
    except KeyError:
        wb.close()
        return []
    out: list[tuple[str, float | None, str | None, str | None, str | None]] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or row[0] is None:
            continue
        a0, a1 = row[0], row[1] if len(row) > 1 else None
        # Sheet data has name in col A and client # in col B (headers are swapped vs body).
        if isinstance(a0, str) and not str(a0).strip().isdigit():
            name = str(a0).strip()
            client_no = float(a1) if isinstance(a1, (int, float)) else None
        elif isinstance(a1, str):
            name = str(a1).strip()
            client_no = float(a0) if isinstance(a0, (int, float)) else None
        else:
            continue
        if not name or name == "اسم العميل":
            continue
        phone = _cell_str(row[4]) if len(row) > 4 else ""
        email = _cell_str(row[5]) if len(row) > 5 else ""
        tax = _cell_str(row[6]) if len(row) > 6 else ""
        out.append(
            (
                name[:150],
                client_no,
                phone if phone and phone != "0000000000" else None,
                email or None,
                tax or None,
            )
        )
    wb.close()
    return out


def persist_block(
    db,
    building_name: str,
    rows: list[UnitRow] | list[ScheduleRow],
    placeholder_owner: Owner,
    dry_run: bool,
    source: str,
) -> tuple[int, int, int]:
    """Returns (buildings_created, units_created, contracts_created)."""
    b_created = u_created = c_created = 0
    if dry_run:
        return (1, len(rows), len(rows))

    b = Building(
        owner_id=placeholder_owner.id,
        name=building_name[:150],
        name_ar=building_name[:150],
        notes=f"استيراد Excel ({source})",
    )
    db.add(b)
    db.flush()
    b_created = 1

    for r in rows:
        if isinstance(r, UnitRow):
            unum = r.unit_number
            rent = r.rent
            tenant_name = r.tenant_name
            phone = r.phone
            pay_method = r.pay_method
            d_start, d_end = r.pay1, r.pay2
            nid_raw = r.national_id_raw
            email = r.email
            electric = r.electric
            water = r.water
        else:
            unum = r.unit_number
            rent = r.rent
            tenant_name = r.tenant_name
            phone = r.phone
            pay_method = r.pay_method
            d_start, d_end = r.start, r.end
            if d_start is None and r.mid is not None:
                d_start = r.mid
            nid_raw = r.national_id_raw
            email = None
            electric = r.electric
            water = r.water

        notes_parts = []
        if water is not None:
            notes_parts.append(f"مياه: {water}")
        if electric:
            notes_parts.append(f"كهرباء: {electric}")
        notes = " | ".join(notes_parts) if notes_parts else None

        unit = Unit(
            building_id=b.id,
            name=f"وحدة {unum}"[:100],
            name_ar=f"وحدة {unum}"[:100],
            number=unum[:50],
            rent_amount=rent,
            electric_invoice=(electric[:50] if electric else None),
            is_available=False,
            notes=notes,
        )
        db.add(unit)
        db.flush()
        u_created += 1

        nid = _national_id(nid_raw, phone, tenant_name)
        tenant = Tenant(
            name=tenant_name,
            name_ar=tenant_name,
            phone=phone,
            national_id=nid,
            email=email,
        )
        db.add(tenant)
        db.flush()

        if d_start is None:
            d_start = date.today()
        if d_end is None or d_end <= d_start:
            try:
                d_end = date(d_start.year + 1, d_start.month, d_start.day)
            except ValueError:
                d_end = date(d_start.year + 1, 12, 31)
        c = Contract(
            unit_id=unit.id,
            tenant_id=tenant.id,
            contract_number=_contract_no("XLS"),
            start_date=d_start,
            end_date=d_end,
            rent_amount=rent,
            payment_cycle=_payment_cycle(pay_method),
            status="active",
            notes=f"استيراد {source}",
        )
        db.add(c)
        c_created += 1

    return b_created, u_created, c_created


def main() -> None:
    ap = argparse.ArgumentParser(description="Migrate property Excel files into Muktasabat DB.")
    ap.add_argument(
        "--data-xlsx",
        default=os.path.expanduser(
            "~/Downloads/بيانات عمائر ادارة الاملاك.xlsx"
        ),
        help="Workbook: بيانات عمائر (العمائر + عملاء)",
    )
    ap.add_argument(
        "--schedule-xlsx",
        default=os.path.expanduser(
            "~/Downloads/جدول عمائر إدارة الأملاك.xlsx"
        ),
        help="Workbook: جدول عمائر (sheet الورقة1)",
    )
    ap.add_argument("--dry-run", action="store_true", help="Parse only; do not write.")
    args = ap.parse_args()

    if not os.path.isfile(args.data_xlsx):
        print("Missing data file:", args.data_xlsx, file=sys.stderr)
        sys.exit(1)
    if not os.path.isfile(args.schedule_xlsx):
        print("Missing schedule file:", args.schedule_xlsx, file=sys.stderr)
        sys.exit(1)

    clients = import_clients(args.data_xlsx)
    blocks_a = list(iter_alomair_units_sheet(args.data_xlsx))
    blocks_b = list(iter_schedule_sheet(args.schedule_xlsx))

    print(f"Clients sheet rows: {len(clients)}")
    print(f"Building blocks (بيانات/العمائر): {len(blocks_a)}")
    print(f"Building blocks (جدول/الورقة1): {len(blocks_b)}")
    if args.dry_run:
        for name, units in blocks_a[:3]:
            print(f"  [عمائر] {name!r} -> {len(units)} units")
        for name, rows in blocks_b[:3]:
            print(f"  [جدول] {name!r} -> {len(rows)} rows")
        print("Dry run complete.")
        return

    with SessionLocal() as db:
        ph = db.scalar(select(Owner).where(Owner.name == "استيراد Excel — مالك افتراضي"))
        if ph is None:
            ph = Owner(
                name="استيراد Excel — مالك افتراضي",
                name_ar="استيراد Excel — مالك افتراضي",
                notes="Created by migrate_property_excel.py for buildings without matched owner",
            )
            db.add(ph)
            db.commit()
            db.refresh(ph)

        for name, _client_no, phone, email, tax in clients:
            exists = db.scalar(select(Owner).where(Owner.name == name))
            if exists:
                continue
            nid = None
            if tax and re.sub(r"\D", "", str(tax)):
                digits = re.sub(r"\D", "", str(tax))[:20]
                if digits:
                    nid = digits
            o = Owner(
                name=name,
                name_ar=name,
                phone=phone,
                email=email,
                national_id=nid,
                notes="استيراد عميل من Excel",
            )
            db.add(o)
        db.commit()

        total_b = total_u = total_c = 0
        for name, units in blocks_a:
            bc, uc, cc = persist_block(db, name, units, ph, False, "بيانات/العمائر")
            total_b += bc
            total_u += uc
            total_c += cc
        db.commit()

        for name, rows in blocks_b:
            bc, uc, cc = persist_block(db, name, rows, ph, False, "جدول/الورقة1")
            total_b += bc
            total_u += uc
            total_c += cc
        db.commit()

        print(f"Done. Buildings={total_b}, units={total_u}, contracts={total_c}")


if __name__ == "__main__":
    main()
